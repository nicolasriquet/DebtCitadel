import logging
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from tempfile import NamedTemporaryFile

from datetime import datetime
import operator
from django.contrib.auth.models import User
from django.conf import settings
from django.contrib import messages
from django.core.exceptions import ValidationError, PermissionDenied
from django.urls import reverse, Resolver404
from django.db.models import Q, Count
from django.http import HttpResponseRedirect, StreamingHttpResponse, HttpResponse, FileResponse, QueryDict
from django.shortcuts import render, get_object_or_404
from django.views.decorators.cache import cache_page
from django.utils import timezone
from time import strftime
from django.contrib.admin.utils import NestedObjects
from django.db import DEFAULT_DB_ALIAS

from dojo.debt_engagement.services import close_debt_engagement, reopen_debt_engagement
from dojo.filters import DebtEngagementFilter, DebtEngagementDirectFilter, DebtEngagementDebtTestFilter
from dojo.forms import CheckForm, \
    UploadThreatForm, RiskAcceptanceForm, NoteForm, DoneForm, \
    EngForm, DebtEngForm, TestForm, DebtTestForm, ReplaceRiskAcceptanceProofForm, AddDebtItemsRiskAcceptanceForm, DeleteDebtEngagementForm, ImportScanForm, \
    CredMappingForm, JIRADebtEngagementForm, JIRAImportScanForm, TypedNoteForm, JIRAProjectForm, \
    EditRiskAcceptanceForm

from dojo.models import Debt_Item, Debt_Context, Debt_Engagement, Debt_Test, \
    Check_List, Test_Import, Debt_Test_Import, Notes, \
    Risk_Acceptance, Development_Environment, Endpoint, \
    Cred_Mapping, Debt_Cred_Mapping, System_Settings, Note_Type, Debt_Context_API_Scan_Configuration
from dojo.tools.factory import get_scan_types_sorted
from dojo.utils import add_error_message_to_response, add_success_message_to_response, get_page_items, add_breadcrumb, handle_uploaded_threat, \
    FileIterWrapper, get_cal_event, Debt_Context_Tab, is_scan_file_too_large, async_delete, \
    get_system_setting, get_setting, redirect_to_return_url_or_else, get_return_url, calculate_grade
from dojo.debt_notifications.helper import create_notification
from dojo.debt_item.views import find_available_notetypes
from functools import reduce
from django.db.models.query import Prefetch, QuerySet
import dojo.jira_link.helper as jira_helper
import dojo.risk_acceptance.helper as ra_helper
from dojo.risk_acceptance.helper import prefetch_for_expiration
from dojo.finding.helper import NOT_ACCEPTED_FINDINGS_QUERY
from dojo.debt_item.helper import NOT_ACCEPTED_DEBT_ITEMS_QUERY
from django.views.decorators.vary import vary_on_cookie
from dojo.authorization.authorization import user_has_permission_or_403
from dojo.authorization.roles_permissions import Permissions
from dojo.debt_context.queries import get_authorized_debt_contexts
from dojo.engagement.queries import get_authorized_engagements
from dojo.debt_engagement.queries import get_authorized_debt_engagements
from dojo.user.queries import get_authorized_users
from dojo.authorization.authorization_decorators import user_is_authorized
from dojo.importers.importer.importer import DojoDefaultImporter as Importer
import dojo.debt_notifications.helper as notifications_helper
from dojo.endpoint.utils import save_endpoints_to_add


logger = logging.getLogger(__name__)


@cache_page(60 * 5)  # cache for 5 minutes
@vary_on_cookie
def debt_engagement_calendar(request):

    if not get_system_setting('enable_calendar'):
        raise Resolver404()

    if 'lead' not in request.GET or '0' in request.GET.getlist('lead'):
        debt_engagements = get_authorized_debt_engagements(Permissions.Debt_Engagement_View)
    else:
        filters = []
        leads = request.GET.getlist('lead', '')
        if '-1' in request.GET.getlist('lead'):
            leads.remove('-1')
            filters.append(Q(lead__isnull=True))
        filters.append(Q(lead__in=leads))
        debt_engagements = get_authorized_debt_engagements(Permissions.Debt_Engagement_View).filter(reduce(operator.or_, filters))

    debt_engagements = debt_engagements.select_related('lead')
    debt_engagements = debt_engagements.prefetch_related('debt_context')

    add_breadcrumb(
        title="Debt_Engagement Calendar", top_level=True, request=request)
    return render(
        request, 'dojo/calendar.html', {
            'caltype': 'debt_engagements',
            'leads': request.GET.getlist('lead', ''),
            'debt_engagements': debt_engagements,
            'users': get_authorized_users(Permissions.Debt_Engagement_View)
        })


def get_filtered_debt_engagements(request, view):

    if view not in ['all', 'active']:
        raise ValidationError(f'View {view} is not allowed')

    debt_engagements = get_authorized_debt_engagements(Permissions.Debt_Engagement_View).order_by('-target_start')

    if view == 'active':
        debt_engagements = debt_engagements.filter(active=True)

    debt_engagements = debt_engagements.select_related('debt_context', 'debt_context__debt_context_type') \
        .prefetch_related('lead', 'tags', 'debt_context__tags')

    if System_Settings.objects.get().enable_jira:
        debt_engagements = debt_engagements.prefetch_related(
            'jira_project__jira_instance',
            'debt_context__jira_project_set__jira_instance'
        )

    debt_engagements = DebtEngagementDirectFilter(request.GET, queryset=debt_engagements)

    return debt_engagements


def get_debt_test_counts(debt_engagements):
    # Get the debt_test counts per debt_engagement. As a separate query, this is much
    # faster than annotating the above `debt_engagements` query.
    debt_engagement_debt_test_counts = {
        debt_test['debt_engagement']: debt_test['debt_test_count']
        for debt_test in Debt_Test.objects.filter(
            debt_engagement__in=debt_engagements
        ).values(
            'debt_engagement'
        ).annotate(
            debt_test_count=Count('debt_engagement')
        )
    }
    return debt_engagement_debt_test_counts


def debt_engagements(request, view):

    if not view:
        view = 'active'

    filtered_debt_engagements = get_filtered_debt_engagements(request, view)

    engs = get_page_items(request, filtered_debt_engagements.qs, 25)
    debt_context_name_words = sorted(get_authorized_debt_contexts(Permissions.Debt_Context_View).values_list('name', flat=True))
    debt_engagement_name_words = sorted(get_authorized_debt_engagements(Permissions.Debt_Engagement_View).values_list('name', flat=True).distinct())

    add_breadcrumb(
        title=f"{view.capitalize()} Debt_Engagements",
        top_level=not len(request.GET),
        request=request)

    return render(
        request, 'dojo/debt_engagement.html', {
            'debt_engagements': engs,
            'debt_engagement_debt_test_counts': get_debt_test_counts(filtered_debt_engagements.qs),
            'filter_form': filtered_debt_engagements.form,
            'debt_context_name_words': debt_context_name_words,
            'debt_engagement_name_words': debt_engagement_name_words,
            'view': view.capitalize(),
        })


def debt_engagements_all(request):

    debt_contexts_with_debt_engagements = get_authorized_debt_contexts(Permissions.Debt_Engagement_View)
    debt_contexts_with_debt_engagements = debt_contexts_with_debt_engagements.filter(~Q(debt_engagement=None)).distinct()

    # count using prefetch instead of just using 'debt_engagement__set_debt_test_debt_test` to avoid loading all debt_test in memory just to count them
    filter_qs = debt_contexts_with_debt_engagements.prefetch_related(
        Prefetch('debt_engagement_set', queryset=Debt_Engagement.objects.all().annotate(debt_test_count=Count('debt_test__id')))
    )

    filter_qs = filter_qs.prefetch_related(
        'debt_engagement_set__tags',
        'debt_context_type',
        'debt_engagement_set__lead',
        'tags',
    )
    if System_Settings.objects.get().enable_jira:
        filter_qs = filter_qs.prefetch_related(
            'debt_engagement_set__jira_project__jira_instance',
            'jira_project_set__jira_instance'
        )

    filtered = DebtEngagementFilter(
        request.GET,
        queryset=filter_qs
    )

    debt_contexts = get_page_items(request, filtered.qs, 25)

    name_words = debt_contexts_with_debt_engagements.values_list('name', flat=True)
    eng_words = get_authorized_debt_engagements(Permissions.Debt_Engagement_View).values_list('name', flat=True).distinct()

    add_breadcrumb(
        title="All Debt_Engagements",
        top_level=not len(request.GET),
        request=request)

    return render(
        request, 'dojo/debt_engagements_all.html', {
            'debt_contexts': debt_contexts,
            'filter_form': filtered.form,
            'name_words': sorted(set(name_words)),
            'eng_words': sorted(set(eng_words)),
        })


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_Edit, 'eid')
def edit_debt_engagement(request, eid):
    debt_engagement = Debt_Engagement.objects.get(pk=eid)
    is_ci_cd = debt_engagement.debt_engagement_type == "CI/CD"
    jira_project_form = None
    jira_epic_form = None
    jira_project = None
    jira_error = False

    if request.method == 'POST':
        form = DebtEngForm(request.POST, instance=debt_engagement, cicd=is_ci_cd, debt_context=debt_engagement.debt_context, user=request.user)
        jira_project = jira_helper.get_jira_project(debt_engagement, use_inheritance=False)

        if form.is_valid():
            # first save debt_engagement details
            new_status = form.cleaned_data.get('status')
            debt_engagement.debt_context = form.cleaned_data.get('debt_context')
            debt_engagement = form.save(commit=False)
            if (new_status == "Cancelled" or new_status == "Completed"):
                debt_engagement.active = False
                create_notification(event='close_debt_engagement',
                        title='Closure of %s' % debt_engagement.name,
                        description='The debt_engagement "%s" was closed' % (debt_engagement.name),
                        debt_engagement=debt_engagement, url=reverse('debt_engagement_all_debt_items', args=(debt_engagement.id, ))),
            else:
                debt_engagement.active = True
            debt_engagement.save()
            form.save_m2m()

            messages.add_message(
                request,
                messages.SUCCESS,
                'Debt Engagement updated successfully.',
                extra_tags='alert-success')

            success, jira_project_form = jira_helper.debt_process_jira_project_form(request, instance=jira_project, target='debt_engagement', debt_engagement=debt_engagement, debt_context=debt_engagement.debt_context)
            error = not success

            success, jira_epic_form = jira_helper.debt_process_jira_epic_form(request, debt_engagement=debt_engagement)
            error = error or not success

            if not error:
                if '_Add Debt_Tests' in request.POST:
                    return HttpResponseRedirect(
                        reverse('add_debt_tests', args=(debt_engagement.id, )))
                else:
                    return HttpResponseRedirect(
                        reverse('view_debt_engagement', args=(debt_engagement.id, )))
        else:
            logger.debug(form.errors)

    else:
        form = DebtEngForm(initial={'debt_context': debt_engagement.debt_context}, instance=debt_engagement, cicd=is_ci_cd, debt_context=debt_engagement.debt_context, user=request.user)

        jira_epic_form = None
        if get_system_setting('enable_jira'):
            jira_project = jira_helper.get_jira_project(debt_engagement, use_inheritance=False)
            jira_project_form = JIRAProjectForm(instance=jira_project, target='debt_engagement', debt_context=debt_engagement.debt_context)
            logger.debug('showing jira-epic-form')
            jira_epic_form = JIRADebtEngagementForm(instance=debt_engagement)

    if is_ci_cd:
        title = 'Edit CI/CD Debt_Engagement'
    else:
        title = 'Edit Interactive Debt_Engagement'

    debt_context_tab = Debt_Context_Tab(debt_engagement.debt_context, title=title, tab="debt_engagements")
    debt_context_tab.setDebtEngagement(debt_engagement)
    return render(request, 'dojo/debt_new_debt_eng.html', {
        'debt_context_tab': debt_context_tab,
        'title': title,
        'form': form,
        'edit': True,
        'jira_epic_form': jira_epic_form,
        'jira_project_form': jira_project_form,
        'debt_engagement': debt_engagement,
    })


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_Delete, 'eid')
def delete_debt_engagement(request, eid):
    debt_engagement = get_object_or_404(Debt_Engagement, pk=eid)
    debt_context = debt_engagement.debt_context
    form = DeleteDebtEngagementForm(instance=debt_engagement)

    if request.method == 'POST':
        if 'id' in request.POST and str(debt_engagement.id) == request.POST['id']:
            form = DeleteDebtEngagementForm(request.POST, instance=debt_engagement)
            if form.is_valid():
                debt_context = debt_engagement.debt_context
                if get_setting("ASYNC_OBJECT_DELETE"):
                    async_del = async_delete()
                    async_del.delete(debt_engagement)
                    message = 'Debt Engagement and relationships will be removed in the background.'
                else:
                    message = 'Debt Engagement and relationships removed.'
                    debt_engagement.delete()
                messages.add_message(
                    request,
                    messages.SUCCESS,
                    message,
                    extra_tags='alert-success')
                create_notification(event='other',
                                    title='Deletion of %s' % debt_engagement.name,
                                    debt_context=debt_context,
                                    description='The debt engagement "%s" was deleted by %s' % (debt_engagement.name, request.user),
                                    url=request.build_absolute_uri(reverse('view_debt_engagements', args=(debt_context.id, ))),
                                    recipients=[debt_engagement.lead],
                                    icon="exclamation-triangle")

                return HttpResponseRedirect(reverse("view_debt_engagements", args=(debt_context.id, )))

    rels = ['Previewing the relationships has been disabled.', '']
    display_preview = get_setting('DELETE_PREVIEW')
    if display_preview:
        collector = NestedObjects(using=DEFAULT_DB_ALIAS)
        collector.collect([debt_engagement])
        rels = collector.nested()

    debt_context_tab = Debt_Context_Tab(debt_context, title="Delete Debt_Engagement", tab="debt_engagements")
    debt_context_tab.setDebtEngagement(debt_engagement)
    return render(request, 'dojo/debt_delete_debt_engagement.html', {
        'debt_context_tab': debt_context_tab,
        'debt_engagement': debt_engagement,
        'form': form,
        'rels': rels,
    })


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_Edit, 'eid')
def copy_debt_engagement(request, eid):
    debt_engagement = get_object_or_404(Debt_Engagement, id=eid)
    debt_context = debt_engagement.debt_context
    form = DoneForm()

    if request.method == 'POST':
        form = DoneForm(request.POST)
        if form.is_valid():
            debt_engagement_copy = debt_engagement.copy()
            calculate_grade(debt_context)
            messages.add_message(
                request,
                messages.SUCCESS,
                'Debt Engagement copied successfully.',
                extra_tags='alert-success')
            create_notification(event='other',
                                title='Copying of %s' % debt_engagement.name,
                                description='The debt_engagement "%s" was copied by %s' % (debt_engagement.name, request.user),
                                debt_context=debt_context,
                                url=request.build_absolute_uri(reverse('view_debt_engagement', args=(debt_engagement_copy.id, ))),
                                recipients=[debt_engagement.lead],
                                icon="exclamation-triangle")
            return redirect_to_return_url_or_else(request, reverse("view_debt_engagements", args=(debt_context.id, )))
        else:
            messages.add_message(
                request,
                messages.ERROR,
                'Unable to copy debt engagement, please try again.',
                extra_tags='alert-danger')

    debt_context_tab = Debt_Context_Tab(debt_context, title="Copy Debt_Engagement", tab="debt_engagements")
    return render(request, 'dojo/copy_object.html', {
        'source': debt_engagement,
        'source_label': 'Debt_Engagement',
        'destination_label': 'Debt_Context',
        'debt_context_tab': debt_context_tab,
        'form': form,
    })


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_View, 'eid')
def view_debt_engagement(request, eid):
    eng = get_object_or_404(Debt_Engagement, id=eid)
    debt_tests = eng.debt_test_set.all().order_by('debt_test_type__name', '-updated')

    default_page_num = 10

    debt_tests_filter = DebtEngagementDebtTestFilter(request.GET, queryset=debt_tests, debt_engagement=eng)
    paged_debt_tests = get_page_items(request, debt_tests_filter.qs, default_page_num)
    # prefetch only after creating the filters to avoid https://code.djangoproject.com/ticket/23771 and https://code.djangoproject.com/ticket/25375
    paged_debt_tests.object_list = prefetch_for_view_debt_tests(paged_debt_tests.object_list)

    debt_context = eng.debt_context
    risks_accepted = eng.risk_acceptance.all().select_related('owner').annotate(accepted_debt_items_count=Count('accepted_debt_items__id'))
    preset_debt_test_type = None
    network = None
    if eng.preset:
        preset_debt_test_type = eng.preset.debt_test_type.all()
        network = eng.preset.network_locations.all()
    system_settings = System_Settings.objects.get()

    jissue = jira_helper.get_jira_issue(eng)
    jira_project = jira_helper.get_jira_project(eng)

    try:
        check = Check_List.objects.get(debt_engagement=eng)
    except:
        check = None
        pass
    notes = eng.notes.all()
    note_type_activation = Note_Type.objects.filter(is_active=True).count()
    if note_type_activation:
        available_note_types = find_available_notetypes(notes)
    form = DoneForm()
    files = eng.files.all()
    if request.method == 'POST':
        user_has_permission_or_403(request.user, eng, Permissions.Note_Add)
        eng.progress = 'check_list'
        eng.save()

        if note_type_activation:
            form = TypedNoteForm(request.POST, available_note_types=available_note_types)
        else:
            form = NoteForm(request.POST)
        if form.is_valid():
            new_note = form.save(commit=False)
            new_note.author = request.user
            new_note.date = timezone.now()
            new_note.save()
            eng.notes.add(new_note)
            if note_type_activation:
                form = TypedNoteForm(available_note_types=available_note_types)
            else:
                form = NoteForm()
            url = request.build_absolute_uri(reverse("view_debt_engagement", args=(eng.id,)))
            title = "Debt_Engagement: %s on %s" % (eng.name, eng.debt_context.name)
            messages.add_message(request,
                                 messages.SUCCESS,
                                 'Note added successfully.',
                                 extra_tags='alert-success')
    else:
        if note_type_activation:
            form = TypedNoteForm(available_note_types=available_note_types)
        else:
            form = NoteForm()

    creds = Debt_Cred_Mapping.objects.filter(
        debt_context=eng.debt_context).select_related('cred_id').order_by('cred_id')
    cred_eng = Debt_Cred_Mapping.objects.filter(
        debt_engagement=eng.id).select_related('cred_id').order_by('cred_id')

    add_breadcrumb(parent=eng, top_level=False, request=request)

    title = ""
    if eng.debt_engagement_type == "CI/CD":
        title = " CI/CD"
    debt_context_tab = Debt_Context_Tab(debt_context, title="View" + title + " Debt_Engagement", tab="debt_engagements")
    debt_context_tab.setDebtEngagement(eng)
    return render(
        request, 'dojo/debt_view_debt_eng.html', {
            'eng': eng,
            'debt_context_tab': debt_context_tab,
            'system_settings': system_settings,
            'debt_tests': paged_debt_tests,
            'filter': debt_tests_filter,
            'check': check,
            'threat': eng.tmodel_path,
            'form': form,
            'notes': notes,
            'files': files,
            'risks_accepted': risks_accepted,
            'jissue': jissue,
            'jira_project': jira_project,
            'creds': creds,
            'cred_eng': cred_eng,
            'network': network,
            'preset_debt_test_type': preset_debt_test_type
        })


def prefetch_for_view_debt_tests(debt_tests):
    prefetched = debt_tests
    if isinstance(debt_tests,
                  QuerySet):  # old code can arrive here with debt_contexts being a list because the query was already executed

        prefetched = prefetched.select_related('lead')
        prefetched = prefetched.prefetch_related('tags', 'debt_test_type', 'notes')
        prefetched = prefetched.annotate(count_debt_items_debt_test_all=Count('debt_item__id', distinct=True))
        prefetched = prefetched.annotate(count_debt_items_debt_test_active=Count('debt_item__id', filter=Q(debt_item__active=True), distinct=True))
        prefetched = prefetched.annotate(count_debt_items_debt_test_active_verified=Count('debt_item__id', filter=Q(debt_item__active=True) & Q(debt_item__verified=True), distinct=True))
        prefetched = prefetched.annotate(count_debt_items_debt_test_mitigated=Count('debt_item__id', filter=Q(debt_item__is_mitigated=True), distinct=True))
        prefetched = prefetched.annotate(count_debt_items_debt_test_dups=Count('debt_item__id', filter=Q(debt_item__duplicate=True), distinct=True))
        prefetched = prefetched.annotate(total_reimport_count=Count('debt_test_import__id', filter=Q(debt_test_import__type=Debt_Test_Import.REIMPORT_TYPE), distinct=True))

    else:
        logger.warning('unable to prefetch because query was already executed')

    return prefetched


@user_is_authorized(Debt_Engagement, Permissions.Debt_Test_Add, 'eid')
def add_debt_tests(request, eid):
    eng = Debt_Engagement.objects.get(id=eid)
    cred_form = CredMappingForm()
    cred_form.fields["cred_user"].queryset = Debt_Cred_Mapping.objects.filter(
        debt_engagement=eng).order_by('cred_id')

    if request.method == 'POST':
        form = DebtTestForm(request.POST, debt_engagement=eng)
        cred_form = CredMappingForm(request.POST)
        cred_form.fields["cred_user"].queryset = Debt_Cred_Mapping.objects.filter(
            debt_engagement=eng).order_by('cred_id')
        if form.is_valid():
            new_debt_test = form.save(commit=False)
            # set default scan_type as it's used in reimport
            new_debt_test.scan_type = new_debt_test.debt_test_type.name
            new_debt_test.debt_engagement = eng
            try:
                new_debt_test.lead = User.objects.get(id=form['lead'].value())
            except:
                new_debt_test.lead = None
                pass

            # Set status to in progress if a debt_test is added
            if eng.status != "In Progress" and eng.active is True:
                eng.status = "In Progress"
                eng.save()

            new_debt_test.save()

            # Save the credential to the debt_test
            if cred_form.is_valid():
                if cred_form.cleaned_data['cred_user']:
                    # Select the credential mapping object from the selected list and only allow if the credential is associated with the debt_context
                    cred_user = Debt_Cred_Mapping.objects.filter(
                        pk=cred_form.cleaned_data['cred_user'].id,
                        debt_engagement=eid).first()

                    new_f = cred_form.save(commit=False)
                    new_f.debt_test = new_debt_test
                    new_f.cred_id = cred_user.cred_id
                    new_f.save()

            messages.add_message(
                request,
                messages.SUCCESS,
                'Debt Test added successfully.',
                extra_tags='alert-success')

            notifications_helper.notify_debt_test_created(new_debt_test)

            if '_Add Another Debt_Test' in request.POST:
                return HttpResponseRedirect(
                    reverse('add_debt_tests', args=(eng.id, )))
            elif '_Add Debt_Items' in request.POST:
                return HttpResponseRedirect(
                    reverse('add_debt_items', args=(new_debt_test.id, )))
            elif '_Finished' in request.POST:
                return HttpResponseRedirect(
                    reverse('view_debt_engagement', args=(eng.id, )))
    else:
        form = DebtTestForm(debt_engagement=eng)
        form.initial['target_start'] = eng.target_start
        form.initial['target_end'] = eng.target_end
        form.initial['lead'] = request.user
    add_breadcrumb(
        parent=eng, title="Add Debt_Tests", top_level=False, request=request)
    debt_context_tab = Debt_Context_Tab(eng.debt_context, title="Add Debt_Tests", tab="debt_engagements")
    debt_context_tab.setDebtEngagement(eng)
    return render(request, 'dojo/debt_add_debt_tests.html', {
        'debt_context_tab': debt_context_tab,
        'form': form,
        'cred_form': cred_form,
        'eid': eid,
        'eng': eng
    })


# Cant use the easy decorator because of the potential for either eid/pid being used
def import_scan_results(request, eid=None, pid=None):
    environment = Development_Environment.objects.filter(name='Development').first()  # If 'Development' was removed, None is used
    debt_engagement = None
    form = ImportScanForm(initial={'environment': environment})
    cred_form = CredMappingForm()
    debt_item_count = 0
    jform = None
    user = request.user

    if eid:
        debt_engagement = get_object_or_404(Debt_Engagement, id=eid)
        debt_engagement_or_debt_context = debt_engagement
        cred_form.fields["cred_user"].queryset = Debt_Cred_Mapping.objects.filter(debt_engagement=debt_engagement).order_by('cred_id')
    elif pid:
        debt_context = get_object_or_404(Debt_Context, id=pid)
        debt_engagement_or_debt_context = debt_context
    else:
        raise Exception('Either Debt_Engagement or Debt_Context has to be provided')

    user_has_permission_or_403(user, debt_engagement_or_debt_context, Permissions.Import_Scan_Result)

    push_all_jira_issues = jira_helper.is_push_all_issues(debt_engagement_or_debt_context)

    if request.method == "POST":
        form = ImportScanForm(request.POST, request.FILES)
        cred_form = CredMappingForm(request.POST)
        cred_form.fields["cred_user"].queryset = Debt_Cred_Mapping.objects.filter(
            debt_engagement=debt_engagement).order_by('cred_id')

        if jira_helper.get_jira_project(debt_engagement_or_debt_context):
            jform = JIRAImportScanForm(request.POST, push_all=push_all_jira_issues, prefix='jiraform')
            logger.debug('jform valid: %s', jform.is_valid())
            logger.debug('jform errors: %s', jform.errors)

        if form.is_valid() and (jform is None or jform.is_valid()):
            scan = request.FILES.get('file', None)
            scan_date = form.cleaned_data['scan_date']
            minimum_severity = form.cleaned_data['minimum_severity']
            activeChoice = form.cleaned_data.get('active', None)
            verifiedChoice = form.cleaned_data.get('verified', None)
            scan_type = request.POST['scan_type']
            tags = form.cleaned_data['tags']
            version = form.cleaned_data['version']
            branch_tag = form.cleaned_data.get('branch_tag', None)
            build_id = form.cleaned_data.get('build_id', None)
            commit_hash = form.cleaned_data.get('commit_hash', None)
            api_scan_configuration = form.cleaned_data.get('api_scan_configuration', None)
            service = form.cleaned_data.get('service', None)
            close_old_debt_items = form.cleaned_data.get('close_old_debt_items', None)
            # close_old_debt_items_debt_contextct_scope is a modifier of close_old_debt_items.
            # If it is selected, close_old_debt_items should also be selected.
            close_old_debt_items_debt_context_scope = form.cleaned_data.get('close_old_debt_items_debt_context_scope', None)
            if close_old_debt_items_debt_context_scope:
                close_old_debt_items = True
            # Will save in the provided environment or in the `Development` one if absent
            environment_id = request.POST.get('environment', 'Development')
            environment = Development_Environment.objects.get(id=environment_id)

            group_by = form.cleaned_data.get('group_by', None)
            create_debt_item_groups_for_all_debt_items = form.cleaned_data['create_debt_item_groups_for_all_debt_items']

            # TODO move to form validation?
            if scan and is_scan_file_too_large(scan):
                messages.add_message(request,
                                     messages.ERROR,
                                     "Report file is too large. Maximum supported size is {} MB".format(settings.SCAN_FILE_MAX_SIZE),
                                     extra_tags='alert-danger')
                return HttpResponseRedirect(reverse('import_scan_results', args=(debt_engagement,)))

            # Allows for a debt_test to be imported with an debt_engagement created on the fly
            if debt_engagement is None:
                debt_engagement = Debt_Engagement()
                debt_engagement.name = "AdHoc Import - " + strftime("%a, %d %b %Y %X", timezone.now().timetuple())
                debt_engagement.threat_model = False
                debt_engagement.api_debt_test = False
                debt_engagement.pen_debt_test = False
                debt_engagement.check_list = False
                debt_engagement.target_start = timezone.now().date()
                debt_engagement.target_end = timezone.now().date()
                debt_engagement.debt_context = debt_context
                debt_engagement.active = True
                debt_engagement.status = 'In Progress'
                debt_engagement.version = version
                debt_engagement.branch_tag = branch_tag
                debt_engagement.build_id = build_id
                debt_engagement.commit_hash = commit_hash
                debt_engagement.save()

            # can't use helper as when push_all_jira_issues is True, the checkbox gets disabled and is always false
            # push_to_jira = jira_helper.is_push_to_jira(new_debt_item, jform.cleaned_data.get('push_to_jira'))
            push_to_jira = push_all_jira_issues or (jform and jform.cleaned_data.get('push_to_jira'))
            error = False

            # Save newly added endpoints
            added_endpoints = save_endpoints_to_add(form.endpoints_to_add_list, debt_engagement.debt_context)

            active = None
            if activeChoice:
                if activeChoice == 'force_to_true':
                    active = True
                elif activeChoice == 'force_to_false':
                    active = False
            verified = None
            if verifiedChoice:
                if verifiedChoice == 'force_to_true':
                    verified = True
                elif verifiedChoice == 'force_to_false':
                    verified = False

            try:
                importer = Importer()
                debt_test, debt_item_count, closed_debt_item_count, _ = importer.import_scan(scan, scan_type, debt_engagement, user, environment, active=active, verified=verified, tags=tags,
                            minimum_severity=minimum_severity, endpoints_to_add=list(form.cleaned_data['endpoints']) + added_endpoints, scan_date=scan_date,
                            version=version, branch_tag=branch_tag, build_id=build_id, commit_hash=commit_hash, push_to_jira=push_to_jira,
                            close_old_debt_items=close_old_debt_items, close_old_debt_items_debt_context_scope=close_old_debt_items_debt_context_scope, group_by=group_by, api_scan_configuration=api_scan_configuration, service=service,
                            create_debt_item_groups_for_all_debt_items=create_debt_item_groups_for_all_debt_items)

                message = f'{scan_type} processed a total of {debt_item_count} debt_items'

                if close_old_debt_items:
                    message = message + ' and closed %d debt_items' % (closed_debt_item_count)

                message = message + "."

                add_success_message_to_response(message)

            except Exception as e:
                logger.exception(e)
                add_error_message_to_response('An exception error occurred during the report import:%s' % str(e))
                error = True

            # Save the credential to the debt_test
            if cred_form.is_valid():
                if cred_form.cleaned_data['cred_user']:
                    # Select the credential mapping object from the selected list and only allow if the credential is associated with the debt_context
                    cred_user = Debt_Cred_Mapping.objects.filter(
                        pk=cred_form.cleaned_data['cred_user'].id,
                        debt_engagement=eid).first()

                    new_f = cred_form.save(commit=False)
                    new_f.debt_test = debt_test
                    new_f.cred_id = cred_user.cred_id
                    new_f.save()

            if not error:
                return HttpResponseRedirect(
                    reverse('view_debt_test', args=(debt_test.id, )))

    debt_context_id = None
    custom_breadcrumb = None
    title = "Import Scan Results"
    if debt_engagement:
        debt_context_tab = Debt_Context_Tab(debt_engagement.debt_context, title=title, tab="debt_engagements")
        debt_context_tab.setDebtEngagement(debt_engagement)
    else:
        custom_breadcrumb = {"", ""}
        debt_context_tab = Debt_Context_Tab(debt_context, title=title, tab="debt_items")

    if jira_helper.get_jira_project(debt_engagement_or_debt_context):
        jform = JIRAImportScanForm(push_all=push_all_jira_issues, prefix='jiraform')

    form.fields['endpoints'].queryset = Endpoint.objects.filter(debt_context__id=debt_context_tab.debt_context.id)
    form.fields['api_scan_configuration'].queryset = Debt_Context_API_Scan_Configuration.objects.filter(debt_context__id=debt_context_tab.debt_context.id)
    return render(request,
        'dojo/import_scan_results.html',
        {'form': form,
         'debt_context_tab': debt_context_tab,
         'debt_engagement_or_debt_context': debt_engagement_or_debt_context,
         'custom_breadcrumb': custom_breadcrumb,
         'title': title,
         'cred_form': cred_form,
         'jform': jform,
         'scan_types': get_scan_types_sorted(),
         })


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_Edit, 'eid')
def close_eng(request, eid):
    eng = Debt_Engagement.objects.get(id=eid)
    close_debt_engagement(eng)
    messages.add_message(
        request,
        messages.SUCCESS,
        'Debt Engagement closed successfully.',
        extra_tags='alert-success')
    create_notification(event='close_debt_engagement',
                        title='Closure of %s' % eng.name,
                        description='The debt_engagement "%s" was closed' % (eng.name),
                        debt_engagement=eng, url=reverse('debt_engagement_all_debt_items', args=(eng.id, ))),
    return HttpResponseRedirect(reverse("view_debt_engagements", args=(eng.debt_context.id, )))


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_Edit, 'eid')
def reopen_eng(request, eid):
    eng = Debt_Engagement.objects.get(id=eid)
    reopen_debt_engagement(eng)
    messages.add_message(
        request,
        messages.SUCCESS,
        'Debt Engagement reopened successfully.',
        extra_tags='alert-success')
    create_notification(event='other',
                        title='Reopening of %s' % eng.name,
                        debt_engagement=eng,
                        description='The debt_engagement "%s" was reopened' % (eng.name),
                        url=reverse('view_debt_engagement', args=(eng.id, ))),
    return HttpResponseRedirect(reverse("view_debt_engagements", args=(eng.debt_context.id, )))


"""
Greg:
status: in debt_contextion
method to complete checklists from the debt_engagement view
"""


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_Edit, 'eid')
def complete_checklist(request, eid):
    eng = get_object_or_404(Debt_Engagement, id=eid)
    try:
        checklist = Check_List.objects.get(debt_engagement=eng)
    except:
        checklist = None
        pass

    add_breadcrumb(
        parent=eng,
        title="Complete checklist",
        top_level=False,
        request=request)
    if request.method == 'POST':
        debt_tests = Debt_Test.objects.filter(debt_engagement=eng)
        debt_items = Debt_Item.objects.filter(debt_test__in=debt_tests).all()
        form = CheckForm(request.POST, instance=checklist, debt_items=debt_items)
        if form.is_valid():
            cl = form.save(commit=False)
            try:
                check_l = Check_List.objects.get(debt_engagement=eng)
                cl.id = check_l.id
                cl.save()
                form.save_m2m()
            except:
                cl.debt_engagement = eng
                cl.save()
                form.save_m2m()
                pass
            messages.add_message(
                request,
                messages.SUCCESS,
                'Checklist saved.',
                extra_tags='alert-success')
            return HttpResponseRedirect(
                reverse('view_debt_engagement', args=(eid, )))
    else:
        debt_tests = Debt_Test.objects.filter(debt_engagement=eng)
        debt_items = Debt_Item.objects.filter(debt_test__in=debt_tests).all()
        form = CheckForm(instance=checklist, debt_items=debt_items)

    debt_context_tab = Debt_Context_Tab(eng.debt_context, title="Checklist", tab="debt_engagements")
    debt_context_tab.setDebtEngagement(eng)
    return render(request, 'dojo/checklist.html', {
        'form': form,
        'debt_context_tab': debt_context_tab,
        'eid': eng.id,
        'debt_items': debt_items,
    })


@user_is_authorized(Debt_Engagement, Permissions.Risk_Acceptance, 'eid')
def add_risk_acceptance(request, eid, fid=None):
    eng = get_object_or_404(Debt_Engagement, id=eid)
    debt_item = None
    if fid:
        debt_item = get_object_or_404(Debt_Item, id=fid)

    if not eng.debt_context.enable_full_risk_acceptance:
        raise PermissionDenied()

    if request.method == 'POST':
        form = RiskAcceptanceForm(request.POST, request.FILES)
        if form.is_valid():
            # first capture notes param as it cannot be saved directly as m2m
            notes = None
            if form.cleaned_data['notes']:
                notes = Notes(
                    entry=form.cleaned_data['notes'],
                    author=request.user,
                    date=timezone.now())
                notes.save()

            del form.cleaned_data['notes']

            try:
                # we sometimes see a weird exception here, but are unable to redebt_contextuce.
                # we add some logging in case it happens
                risk_acceptance = form.save()
            except Exception as e:
                logger.debug(vars(request.POST))
                logger.error(vars(form))
                logger.exception(e)
                raise

            # attach note to risk acceptance object now in database
            if notes:
                risk_acceptance.notes.add(notes)

            eng.risk_acceptance.add(risk_acceptance)

            debt_items = form.cleaned_data['accepted_debt_items']

            risk_acceptance = ra_helper.add_debt_items_to_risk_acceptance(risk_acceptance, debt_items)

            messages.add_message(
                request,
                messages.SUCCESS,
                'Risk acceptance saved.',
                extra_tags='alert-success')

            return redirect_to_return_url_or_else(request, reverse('view_debt_engagement', args=(eid, )))
    else:
        risk_acceptance_title_suggestion = 'Accept: %s' % debt_item
        form = RiskAcceptanceForm(initial={'owner': request.user, 'name': risk_acceptance_title_suggestion})

    debt_item_choices = Debt_Item.objects.filter(duplicate=False, debt_test__debt_engagement=eng).filter(NOT_ACCEPTED_DEBT_ITEMS_QUERY).order_by('title')

    form.fields['accepted_debt_items'].queryset = debt_item_choices
    if fid:
        form.fields['accepted_debt_items'].initial = {fid}
    debt_context_tab = Debt_Context_Tab(eng.debt_context, title="Risk Acceptance", tab="debt_engagements")
    debt_context_tab.setDebtEngagement(eng)

    return render(request, 'dojo/add_risk_acceptance.html', {
                  'eng': eng,
                  'debt_context_tab': debt_context_tab,
                  'form': form
                  })


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_View, 'eid')
def view_risk_acceptance(request, eid, raid):
    return view_edit_risk_acceptance(request, eid=eid, raid=raid, edit_mode=False)


@user_is_authorized(Debt_Engagement, Permissions.Risk_Acceptance, 'eid')
def edit_risk_acceptance(request, eid, raid):
    return view_edit_risk_acceptance(request, eid=eid, raid=raid, edit_mode=True)


# will only be called by view_risk_acceptance and edit_risk_acceptance
def view_edit_risk_acceptance(request, eid, raid, edit_mode=False):
    risk_acceptance = get_object_or_404(Risk_Acceptance, pk=raid)
    eng = get_object_or_404(Debt_Engagement, pk=eid)

    if edit_mode and not eng.debt_context.enable_full_risk_acceptance:
        raise PermissionDenied()

    risk_acceptance_form = None
    errors = False

    if request.method == 'POST':
        # deleting before instantiating the form otherwise django messes up and we end up with an empty path value
        if len(request.FILES) > 0:
            logger.debug('new proof uploaded')
            risk_acceptance.path.delete()

        if 'decision' in request.POST:
            old_expiration_date = risk_acceptance.expiration_date
            risk_acceptance_form = EditRiskAcceptanceForm(request.POST, request.FILES, instance=risk_acceptance)
            errors = errors or not risk_acceptance_form.is_valid()
            if not errors:
                logger.debug('path: %s', risk_acceptance_form.cleaned_data['path'])

                risk_acceptance_form.save()

                if risk_acceptance.expiration_date != old_expiration_date:
                    # risk acceptance was changed, check if risk acceptance needs to be reinstated and debt_items made accepted again
                    ra_helper.reinstate(risk_acceptance, old_expiration_date)

                messages.add_message(
                    request,
                    messages.SUCCESS,
                    'Risk Acceptance saved successfully.',
                    extra_tags='alert-success')

        if 'entry' in request.POST:
            note_form = NoteForm(request.POST)
            errors = errors or not note_form.is_valid()
            if not errors:
                new_note = note_form.save(commit=False)
                new_note.author = request.user
                new_note.date = timezone.now()
                new_note.save()
                risk_acceptance.notes.add(new_note)
                messages.add_message(
                    request,
                    messages.SUCCESS,
                    'Note added successfully.',
                    extra_tags='alert-success')

        if 'delete_note' in request.POST:
            note = get_object_or_404(Notes, pk=request.POST['delete_note_id'])
            if note.author.username == request.user.username:
                risk_acceptance.notes.remove(note)
                note.delete()
                messages.add_message(
                    request,
                    messages.SUCCESS,
                    'Note deleted successfully.',
                    extra_tags='alert-success')
            else:
                messages.add_message(
                    request,
                    messages.ERROR,
                    "Since you are not the note's author, it was not deleted.",
                    extra_tags='alert-danger')

        if 'remove_debt_item' in request.POST:
            debt_item = get_object_or_404(
                Debt_Item, pk=request.POST['remove_debt_item_id'])

            ra_helper.remove_debt_item_from_risk_acceptance(risk_acceptance, debt_item)

            messages.add_message(
                request,
                messages.SUCCESS,
                'Debt_Item removed successfully from risk acceptance.',
                extra_tags='alert-success')

        if 'replace_file' in request.POST:
            replace_form = ReplaceRiskAcceptanceProofForm(
                request.POST, request.FILES, instance=risk_acceptance)

            errors = errors or not replace_form.is_valid()
            if not errors:
                replace_form.save()

                messages.add_message(
                    request,
                    messages.SUCCESS,
                    'New Proof uploaded successfully.',
                    extra_tags='alert-success')
            else:
                logger.error(replace_form.errors)

        if 'add_debt_items' in request.POST:
            add_debt_items_form = AddDebtItemsRiskAcceptanceForm(
                request.POST, request.FILES, instance=risk_acceptance)

            errors = errors or not add_debt_items_form.is_valid()
            if not errors:
                debt_items = add_debt_items_form.cleaned_data['accepted_debt_items']

                ra_helper.add_debt_items_to_risk_acceptance(risk_acceptance, debt_items)

                messages.add_message(
                    request,
                    messages.SUCCESS,
                    'Debt_Item%s added successfully.' % ('s' if len(debt_items) > 1
                                                       else ''),
                    extra_tags='alert-success')

        if not errors:
            logger.debug('redirecting to return_url')
            return redirect_to_return_url_or_else(request, reverse("view_risk_acceptance", args=(eid, raid)))
        else:
            logger.error('errors found')

    else:
        if edit_mode:
            risk_acceptance_form = EditRiskAcceptanceForm(instance=risk_acceptance)

    note_form = NoteForm()
    replace_form = ReplaceRiskAcceptanceProofForm(instance=risk_acceptance)
    add_debt_items_form = AddDebtItemsRiskAcceptanceForm(instance=risk_acceptance)

    accepted_debt_items = risk_acceptance.accepted_debt_items.order_by('numerical_severity')
    fpage = get_page_items(request, accepted_debt_items, 15)

    unaccepted_debt_items = Debt_Item.objects.filter(debt_test__in=eng.debt_test_set.all(), risk_accepted=False) \
        .exclude(id__in=accepted_debt_items).order_by("title")
    add_fpage = get_page_items(request, unaccepted_debt_items, 10, 'apage')
    # on this page we need to add unaccepted debt_items as possible debt_items to add as accepted
    add_debt_items_form.fields[
        "accepted_debt_items"].queryset = add_fpage.object_list

    debt_context_tab = Debt_Context_Tab(eng.debt_context, title="Risk Acceptance", tab="debt_engagements")
    debt_context_tab.setDebtEngagement(eng)
    return render(
        request, 'dojo/view_risk_acceptance.html', {
            'risk_acceptance': risk_acceptance,
            'debt_engagement': eng,
            'debt_context_tab': debt_context_tab,
            'accepted_debt_items': fpage,
            'notes': risk_acceptance.notes.all(),
            'eng': eng,
            'edit_mode': edit_mode,
            'risk_acceptance_form': risk_acceptance_form,
            'note_form': note_form,
            'replace_form': replace_form,
            'add_debt_items_form': add_debt_items_form,
            # 'show_add_debt_items_form': len(unaccepted_debt_items),
            'request': request,
            'add_debt_items': add_fpage,
            'return_url': get_return_url(request),
        })


@user_is_authorized(Debt_Engagement, Permissions.Risk_Acceptance, 'eid')
def expire_risk_acceptance(request, eid, raid):
    risk_acceptance = get_object_or_404(prefetch_for_expiration(Risk_Acceptance.objects.all()), pk=raid)
    eng = get_object_or_404(Debt_Engagement, pk=eid)

    ra_helper.expire_now(risk_acceptance)

    return redirect_to_return_url_or_else(request, reverse("view_risk_acceptance", args=(eid, raid)))


@user_is_authorized(Debt_Engagement, Permissions.Risk_Acceptance, 'eid')
def reinstate_risk_acceptance(request, eid, raid):
    risk_acceptance = get_object_or_404(prefetch_for_expiration(Risk_Acceptance.objects.all()), pk=raid)
    eng = get_object_or_404(Debt_Engagement, pk=eid)

    if not eng.debt_context.enable_full_risk_acceptance:
        raise PermissionDenied()

    ra_helper.reinstate(risk_acceptance, risk_acceptance.expiration_date)

    return redirect_to_return_url_or_else(request, reverse("view_risk_acceptance", args=(eid, raid)))


@user_is_authorized(Debt_Engagement, Permissions.Risk_Acceptance, 'eid')
def delete_risk_acceptance(request, eid, raid):
    risk_acceptance = get_object_or_404(Risk_Acceptance, pk=raid)
    eng = get_object_or_404(Debt_Engagement, pk=eid)

    ra_helper.delete(eng, risk_acceptance)

    messages.add_message(
        request,
        messages.SUCCESS,
        'Risk acceptance deleted successfully.',
        extra_tags='alert-success')
    return HttpResponseRedirect(reverse("view_debt_engagement", args=(eng.id, )))


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_View, 'eid')
def download_risk_acceptance(request, eid, raid):
    import mimetypes

    mimetypes.init()

    risk_acceptance = get_object_or_404(Risk_Acceptance, pk=raid)

    response = StreamingHttpResponse(
        FileIterWrapper(
            open(settings.MEDIA_ROOT + "/" + risk_acceptance.path.name, mode='rb')))
    response['Content-Disposition'] = 'attachment; filename="%s"' \
                                      % risk_acceptance.filename()
    mimetype, encoding = mimetypes.guess_type(risk_acceptance.path.name)
    response['Content-Type'] = mimetype
    return response


"""
Greg
status: in debt_contextion
Upload a threat model at the debt_engagement level. Threat models are stored
under media folder
"""


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_Edit, 'eid')
def upload_threatmodel(request, eid):
    eng = Debt_Engagement.objects.get(id=eid)
    add_breadcrumb(
        parent=eng,
        title="Upload a threat model",
        top_level=False,
        request=request)

    if request.method == 'POST':
        form = UploadThreatForm(request.POST, request.FILES)
        if form.is_valid():
            handle_uploaded_threat(request.FILES['file'], eng)
            eng.progress = 'other'
            eng.threat_model = True
            eng.save()
            messages.add_message(
                request,
                messages.SUCCESS,
                'Threat model saved.',
                extra_tags='alert-success')
            return HttpResponseRedirect(
                reverse('view_debt_engagement', args=(eid, )))
    else:
        form = UploadThreatForm()
    debt_context_tab = Debt_Context_Tab(eng.debt_context, title="Upload Threat Model", tab="debt_engagements")
    return render(request, 'dojo/up_threat.html', {
        'form': form,
        'debt_context_tab': debt_context_tab,
        'eng': eng,
    })


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_View, 'eid')
def view_threatmodel(request, eid):
    eng = get_object_or_404(Debt_Engagement, pk=eid)
    response = FileResponse(open(eng.tmodel_path, 'rb'))
    return response


@user_is_authorized(Debt_Engagement, Permissions.Debt_Engagement_View, 'eid')
def debt_engagement_ics(request, eid):
    eng = get_object_or_404(Debt_Engagement, id=eid)
    start_date = datetime.combine(eng.target_start, datetime.min.time())
    end_date = datetime.combine(eng.target_end, datetime.max.time())
    uid = "dojo_eng_%d_%d" % (eng.id, eng.debt_context.id)
    cal = get_cal_event(
        start_date, end_date,
        "Debt_Engagement: %s (%s)" % (eng.name, eng.debt_context.name),
        "Set aside for debt_engagement %s, on debt_context %s.  Additional detail can be found at %s"
        % (eng.name, eng.debt_context.name,
           request.build_absolute_uri(
               (reverse("view_debt_engagement", args=(eng.id, ))))), uid)
    output = cal.serialize()
    response = HttpResponse(content=output)
    response['Content-Type'] = 'text/calendar'
    response['Content-Disposition'] = 'attachment; filename=%s.ics' % eng.name
    return response


def get_list_index(list, index):
    try:
        element = list[index]
    except Exception as e:
        element = None
    return element


def get_debt_engagements(request):
    url = request.META.get('QUERY_STRING')
    if not url:
        raise ValidationError('Please use the export button when exporting debt_engagements')
    else:
        if url.startswith('url='):
            url = url[4:]

    path_items = list(filter(None, re.split(r'/|\?', url)))

    if not path_items or path_items[0] != 'debt_engagement':
        raise ValidationError('URL is not an debt_engagement view')

    view = query = None
    if get_list_index(path_items, 1) in ['active', 'all']:
        view = get_list_index(path_items, 1)
        query = get_list_index(path_items, 2)
    else:
        view = 'active'
        query = get_list_index(path_items, 1)

    request.GET = QueryDict(query)
    debt_engagements = get_filtered_debt_engagements(request, view).qs
    debt_test_counts = get_debt_test_counts(debt_engagements)

    return debt_engagements, debt_test_counts


def get_excludes():
    return ['is_ci_cd', 'jira_issue', 'jira_project', 'objects', 'unaccepted_open_debt_items']


def get_foreign_keys():
    return ['build_server', 'lead', 'orchestration_engine', 'preset', 'debt_context',
        'report_type', 'requester', 'source_code_management_server']


def csv_export(request):
    debt_engagements, debt_test_counts = get_debt_engagements(request)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=debt_engagements.csv'

    writer = csv.writer(response)

    first_row = True
    for debt_engagement in debt_engagements:
        if first_row:
            fields = []
            for key in dir(debt_engagement):
                if key not in get_excludes() and not callable(getattr(debt_engagement, key)) and not key.startswith('_'):
                    fields.append(key)
            fields.append('debt_tests')

            writer.writerow(fields)

            first_row = False
        if not first_row:
            fields = []
            for key in dir(debt_engagement):
                if key not in get_excludes() and not callable(getattr(debt_engagement, key)) and not key.startswith('_'):
                    value = debt_engagement.__dict__.get(key)
                    if key in get_foreign_keys() and getattr(debt_engagement, key):
                        value = str(getattr(debt_engagement, key))
                    if value and isinstance(value, str):
                        value = value.replace('\n', ' NEWLINE ').replace('\r', '')
                    fields.append(value)
            fields.append(debt_test_counts.get(debt_engagement.id, 0))

            writer.writerow(fields)

    return response


def excel_export(request):
    debt_engagements, debt_test_counts = get_debt_engagements(request)

    workbook = Workbook()
    workbook.iso_dates = True
    worksheet = workbook.active
    worksheet.title = 'Debt_Engagements'

    font_bold = Font(bold=True)

    row_num = 1
    for debt_engagement in debt_engagements:
        if row_num == 1:
            col_num = 1
            for key in dir(debt_engagement):
                if key not in get_excludes() and not callable(getattr(debt_engagement, key)) and not key.startswith('_'):
                    cell = worksheet.cell(row=row_num, column=col_num, value=key)
                    cell.font = font_bold
                    col_num += 1
            cell = worksheet.cell(row=row_num, column=col_num, value='debt_tests')
            cell.font = font_bold
            row_num = 2
        if row_num > 1:
            col_num = 1
            for key in dir(debt_engagement):
                if key not in get_excludes() and not callable(getattr(debt_engagement, key)) and not key.startswith('_'):
                    value = debt_engagement.__dict__.get(key)
                    if key in get_foreign_keys() and getattr(debt_engagement, key):
                        value = str(getattr(debt_engagement, key))
                    if value and isinstance(value, datetime):
                        value = value.replace(tzinfo=None)
                    worksheet.cell(row=row_num, column=col_num, value=value)
                    col_num += 1
            worksheet.cell(row=row_num, column=col_num, value=debt_test_counts.get(debt_engagement.id, 0))
        row_num += 1

    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    response = HttpResponse(
        content=stream,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=debt_engagements.xlsx'
    return response
